# SPDX-FileCopyrightText: 2025 MiromindAI
#
# SPDX-License-Identifier: Apache-2.0


"""
Aggregate benchmark results from multiple runs into a single CSV file.

This script reads benchmark_results.jsonl files from different runs and aggregates
the results into a CSV format with task_id, task_question, ground_truth, and
model_boxed_answer from each run.
"""

import json
from pathlib import Path
from typing import Dict, List
import sys
import pandas as pd
from openpyxl.utils import get_column_letter

# Configuration parameters - modify these as needed
BASE_LOG_DIR = "<your_results_dirs>"

# Specify the results directories to aggregate (relative to BASE_LOG_DIR)
# RESULTS_SUBDIR = "claude-sonnet-4_owl_set"
RESULTS_BASE_DIR = f"{BASE_LOG_DIR}"

# Output Excel file path (also in the shared directory)
OUTPUT_EXCEL_FILENAME = "aggregated_results_0824_gaia.xlsx"
# OUTPUT_EXCEL_FILENAME = f"{os.path.basename(BASE_LOG_DIR)}.xlsx"
OUTPUT_EXCEL_PATH = f"{BASE_LOG_DIR}/{OUTPUT_EXCEL_FILENAME}"

# Font color control switch
# When set to True, all fonts are black; when set to False, different colors are displayed based on pass_at_k_success
ALWAYS_BLACK_FONT = False


def find_benchmark_results_files(base_dir: str) -> List[Path]:
    """
    Find all benchmark_results.jsonl files in the directory structure.

    Args:
        base_dir: Base directory to search in

    Returns:
        List of paths to benchmark_results.jsonl files
    """
    base_path = Path(base_dir)
    if not base_path.exists():
        print(f"Error: Base directory {base_dir} does not exist")
        sys.exit(1)

    results_files = []

    # Look for run_* directories and their benchmark_results.jsonl files
    for item in base_path.rglob("**/benchmark_results.jsonl"):
        # Only include files that are in run_* directories
        if any(parent.name.startswith("run_") for parent in item.parents):
            results_files.append(item)

    if not results_files:
        print(f"Error: No benchmark_results.jsonl files found in {base_dir}")
        sys.exit(1)

    print(f"Found {len(results_files)} benchmark_results.jsonl files:")
    for file_path in sorted(results_files):
        print(f"  {file_path}")

    return sorted(results_files)


def extract_run_number(file_path: Path) -> str:
    """
    Extract run number from the file path.

    Args:
        file_path: Path to the benchmark results file

    Returns:
        Run identifier (e.g., "run_1", "run_2")
    """
    for parent in file_path.parents:
        if parent.name.startswith("run_"):
            return parent.name
    return "unknown_run"


def load_benchmark_results(file_path: Path) -> List[Dict]:
    """
    Load benchmark results from a JSONL file.

    Args:
        file_path: Path to the JSONL file

    Returns:
        List of benchmark result dictionaries
    """
    results = []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if line:
                    try:
                        result = json.loads(line)
                        results.append(result)
                    except json.JSONDecodeError as e:
                        print(
                            f"Warning: Error parsing JSON on line {line_num} in {file_path}: {e}"
                        )
                        continue
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return []

    print(f"Loaded {len(results)} results from {file_path}")
    return results


def aggregate_results(
    results_files: List[Path],
) -> tuple[Dict[str, Dict], List[str], List[str]]:
    """
    Aggregate results from multiple files.

    Args:
        results_files: List of paths to benchmark results files

    Returns:
        Tuple of (aggregated_data, run_ids, task_order_from_run1)
    """
    aggregated = {}
    all_run_ids = set()
    task_order_from_run1 = []

    for file_path in results_files:
        run_id = extract_run_number(file_path)
        all_run_ids.add(run_id)

        results = load_benchmark_results(file_path)

        # If this is run_1, capture the task order
        if run_id == "run_1":
            task_order_from_run1 = [
                result.get("task_id", "")
                for result in results
                if result.get("task_id", "")
            ]
            print(f"Captured task order from run_1: {len(task_order_from_run1)} tasks")

        for result in results:
            task_id = result.get("task_id", "")
            task_question = result.get("task_question", "")
            ground_truth = result.get("ground_truth", "")
            model_boxed_answer = result.get("model_boxed_answer", "")
            pass_at_k_success = result.get("pass_at_k_success", False)

            if not task_id:
                print(f"Warning: Missing task_id in result from {file_path}")
                continue

            if task_id not in aggregated:
                aggregated[task_id] = {
                    "task_id": task_id,
                    "task_question": task_question,
                    "ground_truth": ground_truth,
                    "runs": {},
                }
            else:
                # Verify that task_question and ground_truth are consistent across runs
                if aggregated[task_id]["task_question"] != task_question:
                    print(f"Warning: Inconsistent task_question for task_id {task_id}")
                if aggregated[task_id]["ground_truth"] != ground_truth:
                    print(f"Warning: Inconsistent ground_truth for task_id {task_id}")

            # Store the model_boxed_answer and pass_at_k_success for this run
            aggregated[task_id]["runs"][run_id] = {
                "model_boxed_answer": model_boxed_answer,
                "pass_at_k_success": pass_at_k_success,
            }

    print(
        f"Aggregated results for {len(aggregated)} unique tasks across {len(all_run_ids)} runs"
    )
    print(f"Run IDs found: {sorted(all_run_ids)}")

    return aggregated, sorted(all_run_ids), task_order_from_run1


def write_excel(
    aggregated_data: Dict[str, Dict],
    run_ids: List[str],
    task_order: List[str],
    output_path: str,
):
    """
    Write aggregated data to Excel file with conditional formatting.

    Args:
        aggregated_data: Aggregated benchmark results
        run_ids: List of run identifiers
        task_order: List of task_ids in the order from run_1
        output_path: Path to output Excel file
    """
    try:
        # Prepare data for DataFrame
        data_rows = []
        formatting_info = []  # Store formatting information

        # Use task_order from run_1 to maintain the same order
        for row_idx, task_id in enumerate(task_order):
            if task_id not in aggregated_data:
                print(
                    f"Warning: task_id {task_id} from run_1 not found in aggregated data"
                )
                continue

            task_data = aggregated_data[task_id]

            row = {
                "task_id": task_data["task_id"],
                "task_question": task_data["task_question"],
                "ground_truth": task_data["ground_truth"],
            }

            row_formatting = {
                "row_idx": row_idx + 2,
                "runs": {},
            }  # +2 because Excel is 1-indexed and we have headers

            # Add model answers for each run
            for run_id in run_ids:
                run_data = task_data["runs"].get(run_id, {})
                if isinstance(run_data, dict):
                    answer = run_data.get("model_boxed_answer", "")
                    pass_at_k_success = run_data.get("pass_at_k_success", False)
                else:
                    # Backward compatibility for old format
                    answer = run_data
                    pass_at_k_success = False

                # If answer is correct (pass_at_k_success=True) and ALWAYS_BLACK_FONT is False, leave blank; otherwise show the answer
                if ALWAYS_BLACK_FONT:
                    display_answer = (
                        answer  # Always show the answer when ALWAYS_BLACK_FONT is True
                    )
                else:
                    display_answer = (
                        "" if pass_at_k_success else answer
                    )  # Original logic
                row[f"model_boxed_answer_{run_id}"] = display_answer
                row_formatting["runs"][run_id] = pass_at_k_success

            data_rows.append(row)
            formatting_info.append(row_formatting)

        # Calculate accuracy stats based on pass_at_k_success field
        accuracy_stats = []
        for idx, run_id in enumerate(run_ids):
            successes = 0
            total_tasks = len(data_rows)

            # Count based on pass_at_k_success from formatting_info
            for fmt_info in formatting_info:
                pass_at_k_success = fmt_info["runs"].get(run_id, False)
                if pass_at_k_success:
                    successes += 1

            accuracy = successes / total_tasks if total_tasks > 0 else 0
            accuracy_stats.append(
                {
                    "run_id": run_id,
                    "successes": successes,
                    "total": total_tasks,
                    "accuracy": accuracy,
                }
            )

        # Add accuracy stats to the data rows for inclusion in Excel
        # Important: Use only plain text to avoid any formula interpretation
        summary_rows = []
        summary_rows.append(
            ["Accuracy Statistics", "", "", ""]
        )  # Remove "===" which might be interpreted as formula

        for stat in accuracy_stats:
            summary_rows.append(
                [
                    f"{stat['run_id']} Accuracy",  # Remove colon which might cause issues
                    f"{stat['accuracy']:.2%}",
                    f"{stat['successes']} out of {stat['total']}",  # Based on pass_at_k_success
                    "",
                ]
            )

        summary_rows.append(
            ["Total Tasks", str(len(data_rows)), "", ""]
        )  # Convert to string
        summary_rows.append(["Number of Runs", str(len(run_ids)), "", ""])

        # Create initial DataFrame to get column names
        df = pd.DataFrame(data_rows)

        # Create summary rows with proper column mapping
        summary_data = []
        for row in summary_rows:
            summary_dict = {}
            col_names = list(df.columns)
            for i, value in enumerate(row):
                if i < len(col_names):
                    summary_dict[col_names[i]] = value
            summary_data.append(summary_dict)

        # Combine main data with summary
        all_data = data_rows + [{}] + summary_data  # Add empty row as separator
        df_final = pd.DataFrame(all_data)

        # Write to Excel using the safest possible method
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font

            # Create a new workbook from scratch to avoid any pandas-related issues
            wb = Workbook()
            ws = wb.active
            ws.title = "Aggregated Results"

            # Get column headers
            headers = list(df_final.columns)

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Write data rows
            for row_idx, (_, row) in enumerate(df_final.iterrows(), 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row[header]
                    if pd.isna(value):
                        value = ""
                    else:
                        value = str(value)  # Convert everything to string

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Apply text format to avoid number interpretation issues
                    cell.number_format = "@"

                    # Apply conditional formatting for model answer columns
                    if (
                        col_idx >= 4 and row_idx <= len(data_rows) + 1
                    ):  # Only for data rows, not summary
                        # Find corresponding formatting info
                        data_row_idx = row_idx - 2  # Convert to 0-based data index
                        if data_row_idx < len(formatting_info):
                            fmt_info = formatting_info[data_row_idx]
                            run_idx = col_idx - 4  # Convert to run index
                            if run_idx < len(run_ids):
                                run_id = run_ids[run_idx]
                                pass_at_k_success = fmt_info["runs"].get(run_id, False)

                                # Apply font color based on ALWAYS_BLACK_FONT setting
                                if ALWAYS_BLACK_FONT:
                                    cell.font = Font(color="000000")  # Always black
                                else:
                                    if pass_at_k_success:
                                        cell.font = Font(color="808080")  # Light gray
                                    else:
                                        cell.font = Font(color="8B0000")  # Dark red

            # Set column widths to 25
            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 25

            # Set row height to 20
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20

            # Disable error checking to remove green triangles
            ws.ignore_errors = True

            # Save the workbook
            wb.save(output_path)
            wb.close()

            print("Successfully created Excel file with custom formatting")

        except Exception as e:
            print(f"Error creating Excel file: {e}")
            # Ultimate fallback: basic pandas save
            df_final.to_excel(
                output_path,
                sheet_name="Aggregated Results",
                index=False,
                engine="openpyxl",
            )

        print(f"Successfully wrote aggregated results to {output_path}")
        print(
            f"Excel file contains {len(data_rows)} tasks with answers from {len(run_ids)} runs"
        )
        print("Task order matches run_1 order")
        if ALWAYS_BLACK_FONT:
            print("Applied font color: Always black (ALWAYS_BLACK_FONT=True)")
            print(
                "Content display: Always show model answers regardless of pass_at_k_success"
            )
        else:
            print(
                "Applied conditional formatting: pass_at_k_success=True (light gray), False (dark red)"
            )
            print("Content display: Hide answers when pass_at_k_success=True")
        print("Added accuracy calculation formulas at the bottom")

    except Exception as e:
        print(f"Error writing Excel file {output_path}: {e}")
        sys.exit(1)


def main():
    """Main function to orchestrate the aggregation process."""
    print("=== Benchmark Results Aggregation Script ===")
    print(f"Looking for results in: {RESULTS_BASE_DIR}")
    print(f"Output Excel file will be saved to: {OUTPUT_EXCEL_PATH}")
    print()

    # Find all benchmark results files
    results_files = find_benchmark_results_files(RESULTS_BASE_DIR)
    print()

    # Aggregate results from all files
    aggregated_data, run_ids, task_order = aggregate_results(results_files)
    print()

    # Write to Excel
    write_excel(aggregated_data, run_ids, task_order, OUTPUT_EXCEL_PATH)
    print()

    # Summary statistics
    print("=== Summary ===")
    print(f"Total unique tasks: {len(aggregated_data)}")
    print(f"Total runs processed: {len(run_ids)}")
    print(f"Runs: {', '.join(run_ids)}")
    print("Task order preserved from: run_1")

    # Check for missing data
    missing_count = 0
    for task_id, task_data in aggregated_data.items():
        for run_id in run_ids:
            run_data = task_data["runs"].get(run_id, {})
            if not run_data or (
                isinstance(run_data, dict)
                and not run_data.get("model_boxed_answer", "")
            ):
                missing_count += 1

    if missing_count > 0:
        print(f"Warning: {missing_count} missing model answers detected")
    else:
        print("All tasks have answers from all runs")

    print("Aggregation completed successfully!")


if __name__ == "__main__":
    main()
